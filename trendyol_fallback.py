import os, json, io, hashlib
from datetime import datetime

LOG_FILE='madmext_logs.json'

DATA_KEYS=['urun','magaza','influencer','meta','urun_detay','influencer_detay']

TYPE_LABELS={
    'urun':'Ürün Reklamları',
    'urun_detay':'Ürün Detay',
    'magaza':'Mağaza Reklamları',
    'influencer':'Influencer Reklamları',
    'influencer_detay':'Influencer Detay',
    'meta':'Meta Trendyol',
    'influencer_user_summary':'Influencer Özet'
}

def _read():
    try:
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE,'r',encoding='utf-8') as f:return json.load(f)
    except Exception: pass
    return {}

def _write(d):
    with open(LOG_FILE,'w',encoding='utf-8') as f:json.dump(d,f,ensure_ascii=False,indent=2)

def _safe(v): return '' if v is None else str(v).strip()
def _slug(v):
    s=_safe(v).lower().replace('ı','i').replace('ğ','g').replace('ü','u').replace('ş','s').replace('ö','o').replace('ç','c')
    return ' '.join(s.split())
def _norm(v): return _slug(v)
def _day(v): return _safe(v)[:10]
def _num(v):
    try:
        s=_safe(v).replace('₺','').replace('+ KDV','').replace('KDV','').strip()
        return float(s.replace('.','').replace(',','.'))
    except Exception:return 0

def _checksum(data): return hashlib.sha256(data or b'').hexdigest()

def _score(r):
    return _num(r.get('harcama'))+_num(r.get('toplam_ciro') or r.get('ciro'))+_num(r.get('toplam_satis') or r.get('satis'))+_num(r.get('tiklanma'))+_num(r.get('ziyaret'))+_num(r.get('goruntulenme'))+_num(r.get('roas'))

def _key(t,r):
    if t=='urun_detay': return _norm(r.get('kampanya'))+'|'+_norm(r.get('content_id') or r.get('model') or r.get('urun_adi'))
    if t=='influencer_detay': return _norm(r.get('kampanya'))+'|'+_norm(r.get('kullanici'))+'|'+_norm(r.get('urun'))+'|'+_day(r.get('tarih'))
    return _norm(r.get('ad'))+'|'+_day(r.get('baslangic'))

def _dedupe(t,rows):
    m={}
    for r in rows or []:
        k=_key(t,r)
        if not k or k.count('|')==len(k): continue
        old=m.get(k)
        if old is None:
            n=dict(r); n['_dedupe_key']=k; m[k]=n
        elif _score(r)>=_score(old):
            n=dict(old); n.update(r); n['_dedupe_key']=k; m[k]=n
    return list(m.values())

def _headers(ws): return [_safe(c.value) for c in ws[1]]
def _hset(ws): return set(_slug(x) for x in _headers(ws))
def _idx(headers,names):
    sl=[_slug(h) for h in headers]
    for n in names:
        if _slug(n) in sl:return sl.index(_slug(n))
    return -1

def _detect(ws):
    h=_hset(ws)
    if 'kullanici ismi' in h and 'urun' in h and 'tarih' in h:return 'influencer_detay'
    if 'kullanici ismi' in h and 'link ziyareti' in h and 'ciro' in h:return 'influencer_user_summary'
    if 'reklam adi' in h and 'reklam statusu' in h and ('link ziyareti' in h or 'satis adedi' in h):return 'influencer'
    if 'content id' in h or 'model kodu' in h:return 'urun_detay'
    if 'reklam adi' in h and 'harcama getirisi' in h and 'tbm teklifi' in h:return 'urun'
    if 'reklam adi' in h and 'butce tipi' in h:return 'influencer'
    if 'reklam adi' in h and 'toplam butce' in h and 'reklam cirosu' in h:return 'meta'
    if 'reklam adi' in h and 'harcanan butce' in h:return 'magaza'
    if ws.max_column>=20:return 'urun'
    return 'unknown'

def _cols(t):
    return {
      'urun':['ad','statu','baslangic','bitis','urun_adedi','content_ids','toplam_butce','gunluk_butce','kalan_butce','harcama','tbm_teklifi','gerceklesen_tbm','tiklanma','goruntulenme','dogrudan_satis','dolayli_satis','toplam_satis','dogrudan_ciro','dolayli_ciro','toplam_ciro','roas'],
      'urun_detay':['kampanya','urun_adi','content_id','model','harcama','goruntulenme','tiklanma','ctr','dogrudan_satis','dolayli_satis','toplam_satis','dogrudan_ciro','dolayli_ciro','toplam_ciro','roas'],
      'magaza':['ad','statu','baslangic','bitis','harcama','goruntulenme','tiklanma','toplam_satis','toplam_ciro','roas'],
      'influencer':['ad','statu','baslangic','bitis','butce_tipi','odeme','ziyaret','satis','ciro','paylasim'],
      'influencer_detay':['kampanya','kullanici','urun','tarih','ciro','satis','ziyaret'],
      'meta':['ad','statu','baslangic','bitis','toplam_butce','harcama','goruntulenme','tiklanma','satis','ciro','roas']
    }.get(t,[])

def _rows(ws,t,kampanya=''):
    if t=='influencer':
        headers=_headers(ws); out=[]
        ix_ad=_idx(headers,['Reklam Adı','Reklam Adi']); ix_st=_idx(headers,['Reklam Statüsü','Reklam Statusu','Reklam Statu'])
        ix_ba=_idx(headers,['Başlangıç Tarihi','Baslangic Tarihi']); ix_bi=_idx(headers,['Bitiş Tarihi','Bitis Tarihi'])
        ix_od=_idx(headers,['Ödenecek Tutar','Odenecek Tutar']); ix_ko=_idx(headers,['Komisyon']); ix_bo=_idx(headers,['Bonus'])
        ix_zi=_idx(headers,['Link Ziyareti']); ix_sa=_idx(headers,['Satış Adedi','Satis Adedi']); ix_ci=_idx(headers,['Reklam Cirosu','Ciro']); ix_pa=_idx(headers,['Paylaşım Sayısı','Paylasim Sayisi'])
        for row in ws.iter_rows(min_row=2,values_only=True):
            if not row or ix_ad<0 or not row[ix_ad]:continue
            out.append({'ad':_safe(row[ix_ad]),'statu':_safe(row[ix_st]) if ix_st>=0 else '','baslangic':_safe(row[ix_ba]) if ix_ba>=0 else '','bitis':_safe(row[ix_bi]) if ix_bi>=0 else '','butce_tipi':_safe(row[ix_bo]) if ix_bo>=0 else '','odeme':_safe(row[ix_od]) if ix_od>=0 else (_safe(row[ix_ko]) if ix_ko>=0 else ''),'ziyaret':_safe(row[ix_zi]) if ix_zi>=0 else '','satis':_safe(row[ix_sa]) if ix_sa>=0 else '','ciro':_safe(row[ix_ci]) if ix_ci>=0 else '','paylasim':_safe(row[ix_pa]) if ix_pa>=0 else ''})
        return out
    if t=='influencer_user_summary':
        headers=_headers(ws); out=[]
        ix_u=_idx(headers,['Kullanıcı İsmi','Kullanici Ismi']); ix_z=_idx(headers,['Link Ziyareti']); ix_c=_idx(headers,['Ciro']); ix_s=_idx(headers,['Satış Adedi','Satis Adedi'])
        ix_y=_idx(headers,['Yeni Müşteri','Yeni Musteri'])
        for row in ws.iter_rows(min_row=2,values_only=True):
            if not row or ix_u<0 or not row[ix_u]:continue
            out.append({'ad':_safe(row[ix_u]),'statu':'Influencer Bazli Satis','baslangic':'','bitis':'','butce_tipi':'Influencer','odeme':'','ziyaret':_safe(row[ix_z]) if ix_z>=0 else '','satis':_safe(row[ix_s]) if ix_s>=0 else '','ciro':_safe(row[ix_c]) if ix_c>=0 else '','paylasim':_safe(row[ix_y]) if ix_y>=0 else ''})
        return out
    if t=='influencer_detay':
        headers=_headers(ws); out=[]
        ix_u=_idx(headers,['Kullanıcı İsmi','Kullanici Ismi']); ix_p=_idx(headers,['Ürün','Urun']); ix_t=_idx(headers,['Tarih'])
        ix_c=_idx(headers,['Ciro']); ix_s=_idx(headers,['Satış Adedi','Satis Adedi']); ix_z=_idx(headers,['Link Ziyareti',"Influencer'ın Tarih'te getirdiği Link Ziyareti"])
        for row in ws.iter_rows(min_row=2,values_only=True):
            if not row or ix_u<0 or ix_t<0 or not row[ix_u]:continue
            out.append({'kampanya':kampanya or 'Influencer-TR','kullanici':_safe(row[ix_u]),'urun':_safe(row[ix_p]) if ix_p>=0 else '','tarih':_safe(row[ix_t]),'ciro':_safe(row[ix_c]) if ix_c>=0 else '','satis':_safe(row[ix_s]) if ix_s>=0 else '','ziyaret':_safe(row[ix_z]) if ix_z>=0 else ''})
        return out
    cols=_cols(t); out=[]
    for row in ws.iter_rows(min_row=2,values_only=True):
        if not row or not row[0]:continue
        vals=list(row)+['']*30
        if t=='urun_detay':vals=[ws.title]+vals
        out.append(dict(zip(cols,[_safe(v) for v in vals[:len(cols)]])))
    return out

def _parse_wb(wb):
    parsed={'urun':[],'magaza':[],'influencer':[],'meta':[],'urun_detay':[],'influencer_detay':[]}
    detected=[]
    for ws in wb.worksheets:
        t=_detect(ws); detected.append(t)
        if t=='unknown':continue
        if t=='influencer_detay': parsed['influencer_detay'].extend(_rows(ws,t,'Influencer-TR'))
        elif t=='influencer_user_summary': parsed['influencer'].extend(_rows(ws,t))
        else:
            key='influencer' if t=='influencer' else t
            parsed.setdefault(key,[]); parsed[key].extend(_rows(ws,t))
    return parsed,detected

def _counts(d):
    return {k:len(_dedupe(k,(d or {}).get(k,[]))) for k in DATA_KEYS}

def install(app):
    from flask import request,jsonify,session
    @app.before_request
    def ty_no_db_fallback():
        if os.environ.get('DATABASE_URL'):return None
        if request.path=='/trendyol/data' and request.method=='GET':
            logs=_read(); d=(logs.get('trendyol_fallback') or {})
            cleaned={k:_dedupe(k,d.get(k,[])) for k in DATA_KEYS}
            if cleaned!=d: logs['trendyol_fallback']=cleaned; _write(logs)
            return jsonify(cleaned)
        if request.path=='/trendyol/stats' and request.method=='GET':
            d=(_read().get('trendyol_fallback') or {})
            stats=_counts(d)
            uploads=_read().get('trendyol_uploads',[])
            stats['uploads']=len(uploads)
            stats['last_upload']=uploads[0] if uploads else None
            return jsonify(stats)
        if request.path=='/trendyol/uploads' and request.method=='GET':
            logs=_read(); uploads=logs.get('trendyol_uploads',[])
            return jsonify({'uploads':uploads[:200], 'stats':_counts(logs.get('trendyol_fallback') or {})})
        if request.path=='/trendyol/upload' and request.method=='POST':
            f=request.files.get('file')
            if not f:return jsonify({'error':'Dosya bulunamadi','success':False})
            try:
                import openpyxl
                raw=f.read(); ch=_checksum(raw); filename=getattr(f,'filename','') or 'trendyol.xlsx'
                logs=_read(); uploads=logs.setdefault('trendyol_uploads',[])
                existing=next((u for u in uploads if u.get('checksum')==ch),None)
                if existing:
                    existing['last_seen_at']=datetime.utcnow().isoformat()
                    existing['duplicate_seen_count']=int(existing.get('duplicate_seen_count') or 0)+1
                    logs.setdefault('actionLog',[]).insert(0,{'type':'trendyol_duplicate_upload','file':filename,'checksum':ch[:12],'serverTime':datetime.utcnow().isoformat()})
                    logs['actionLog']=logs.get('actionLog',[])[:500]; _write(logs)
                    return jsonify({'success':True,'duplicate':True,'count':0,'type':existing.get('dosya_tipi','Tekrar Yüklenen Dosya'),'tab':existing.get('tab','urun'),'message':'Bu dosya daha önce yüklenmiş. Veri çoğaltılmadı.','upload':existing})
                wb=openpyxl.load_workbook(io.BytesIO(raw))
                parsed,detected=_parse_wb(wb)
                if not any(parsed[k] for k in parsed):return jsonify({'error':'Taninmayan Trendyol Excel raporu','success':False,'detected':detected})
                store=logs.setdefault('trendyol_fallback',{'urun':[],'magaza':[],'influencer':[],'meta':[],'urun_detay':[],'influencer_detay':[]})
                before=_counts(store); incoming=0; after_counts={}
                for k,rows in parsed.items():
                    if not rows:continue
                    incoming+=len(rows)
                    store.setdefault(k,[])
                    store[k]=_dedupe(k,store[k]+rows)[-5000:]
                after=_counts(store)
                for k in DATA_KEYS: after_counts[k]=after.get(k,0)-before.get(k,0)
                main_type=next((x for x in detected if x!='unknown'), 'unknown')
                label='Influencer Reklamlari' if any(x in detected for x in ['influencer','influencer_detay','influencer_user_summary']) else ','.join([TYPE_LABELS.get(x,x) for x in detected if x!='unknown'])
                upload={
                    'upload_id': ch[:16],
                    'dosya_adi': filename,
                    'dosya_tipi': label or TYPE_LABELS.get(main_type,main_type),
                    'tab': 'influencer' if any(x in detected for x in ['influencer','influencer_detay','influencer_user_summary']) else main_type,
                    'yukleme_tarihi': datetime.utcnow().isoformat(),
                    'satir_sayisi': incoming,
                    'tekil_eklenen': sum(v for v in after_counts.values() if v>0),
                    'checksum': ch,
                    'checksum_short': ch[:12],
                    'yukleyen_kullanici': session.get('user_email',''),
                    'detected': detected,
                    'changes': after_counts,
                    'duplicate_seen_count': 0
                }
                uploads.insert(0,upload); logs['trendyol_uploads']=uploads[:300]
                logs.setdefault('actionLog',[]).insert(0,{'type':'trendyol_upload_fallback','detected':detected,'incoming':incoming,'unique_added':upload['tekil_eklenen'],'file':filename,'checksum':ch[:12],'serverTime':datetime.utcnow().isoformat()})
                logs['actionLog']=logs.get('actionLog',[])[:500]; _write(logs)
                return jsonify({'success':True,'count':incoming,'unique_added':upload['tekil_eklenen'],'type':upload['dosya_tipi'],'tab':upload['tab'],'fallback':True,'upload':upload})
            except Exception as e:return jsonify({'error':str(e),'success':False})
        return None
