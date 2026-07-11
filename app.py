from flask import Flask, request, jsonify, send_from_directory, session
from flask_cors import CORS
import requests
import os
import json
import threading
import hashlib
import hmac
import ipaddress
import re
import socket
from urllib.parse import urlsplit
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
try:
    import psycopg2
    import psycopg2.extras
    HAS_PG = True
except ImportError:
    HAS_PG = False
app = Flask(__name__, static_folder='.')

SECRET_KEY = os.environ.get('SECRET_KEY', '').strip()
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', '').strip()
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', '').strip()

if not SECRET_KEY:
    raise RuntimeError('SECRET_KEY ortam değişkeni zorunlu. Railway Variables içine SECRET_KEY ekleyin.')
if not ADMIN_EMAIL:
    raise RuntimeError('ADMIN_EMAIL ortam değişkeni zorunlu. Railway Variables içine ADMIN_EMAIL ekleyin.')
if not ADMIN_PASSWORD:
    raise RuntimeError('ADMIN_PASSWORD ortam değişkeni zorunlu. Railway Variables içine ADMIN_PASSWORD ekleyin.')

app.secret_key = SECRET_KEY
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=os.environ.get('SESSION_COOKIE_SECURE', 'true').lower() != 'false',
    SESSION_COOKIE_SAMESITE='Lax',
)
from datetime import timedelta
app.permanent_session_lifetime = timedelta(days=30)
_cors_origins = [x.strip() for x in os.environ.get(
    'ALLOWED_ORIGINS', 'https://web-production-e5865.up.railway.app'
).split(',') if x.strip()]
CORS(app, supports_credentials=True, origins=_cors_origins)
META_TOKEN = os.environ.get('META_TOKEN', '')
ANTHROPIC_KEY = os.environ.get('ANTHROPIC_KEY', '')
PSI_KEY = os.environ.get('PSI_KEY', '')  # Google PageSpeed Insights API key (opsiyonel ama onerilen)
GA4_PROPERTY_ID = os.environ.get('GA4_PROPERTY_ID', '')
GA4_REFRESH_TOKEN = os.environ.get('GA4_REFRESH_TOKEN', '')
GA4_CLIENT_ID = os.environ.get('GA4_CLIENT_ID', '')
GA4_CLIENT_SECRET = os.environ.get('GA4_CLIENT_SECRET', '')
GADS_DEVELOPER_TOKEN = os.environ.get('GADS_DEVELOPER_TOKEN', '')
GADS_CUSTOMER_ID = os.environ.get('GADS_CUSTOMER_ID', '')
GADS_LOGIN_CUSTOMER_ID = os.environ.get('GADS_LOGIN_CUSTOMER_ID', '')
GADS_CLIENT_ID = os.environ.get('GADS_CLIENT_ID', '') or GA4_CLIENT_ID
GADS_CLIENT_SECRET = os.environ.get('GADS_CLIENT_SECRET', '') or GA4_CLIENT_SECRET
GADS_REFRESH_TOKEN = os.environ.get('GADS_REFRESH_TOKEN', '') or GA4_REFRESH_TOKEN
LOG_FILE = 'madmext_logs.json'
log_lock = threading.Lock()
_users_cache = None
_users_lock = threading.Lock()
_ga4_token_cache = {'token': None, 'expires_at': 0}
_ga4_token_lock = threading.Lock()
def read_logs():
    try:
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except:
        pass
    return {'budgetLog': [], 'taskLog': [], 'actionLog': []}
def write_logs(data):
    with log_lock:
        with open(LOG_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
def get_ga4_token():
    import time
    with _ga4_token_lock:
        now = time.time()
        if _ga4_token_cache['token'] and now < _ga4_token_cache['expires_at']:
            return _ga4_token_cache['token']
        try:
            r = requests.post('https://oauth2.googleapis.com/token', data={
                'client_id': GA4_CLIENT_ID,
                'client_secret': GA4_CLIENT_SECRET,
                'refresh_token': GA4_REFRESH_TOKEN,
                'grant_type': 'refresh_token'
            }, timeout=10)
            token = r.json().get('access_token')
            if token:
                _ga4_token_cache['token'] = token
                _ga4_token_cache['expires_at'] = now + 2900
            return token
        except Exception as e:
            print('GA4 token hata:', e)
            return None
def get_db():
    url = os.environ.get('DATABASE_URL','')
    if not url or not HAS_PG: return None
    try:
        if url.startswith('postgres://'): url = url.replace('postgres://','postgresql://',1)
        return psycopg2.connect(url, sslmode='require')
    except Exception as e:
        print('DB error:', e); return None
def init_db():
    conn = get_db()
    if not conn: return
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS mx_users (
                email TEXT PRIMARY KEY,
                name TEXT,
                password_hash TEXT NOT NULL,
                role TEXT DEFAULT 'viewer',
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        conn.commit(); cur.close(); conn.close()
    except Exception as e: print('init_db:', e)
def db_get_users():
    conn = get_db()
    if not conn: return None
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT email,name,password_hash,role FROM mx_users ORDER BY created_at')
        rows = [dict(r) for r in cur.fetchall()]
        cur.close(); conn.close()
        return rows
    except Exception as e: print('db_get_users:', e); return None
def db_upsert_user(email, name, pw_hash, role):
    conn = get_db()
    if not conn: return False
    try:
        cur = conn.cursor()
        cur.execute(
            'INSERT INTO mx_users(email,name,password_hash,role) VALUES(%s,%s,%s,%s) ON CONFLICT(email) DO UPDATE SET name=%s,role=%s',
            (email,name,pw_hash,role,name,role)
        )
        conn.commit(); cur.close(); conn.close(); return True
    except Exception as e: print('db_upsert:', e); return False
def db_delete(email):
    conn = get_db()
    if not conn: return False
    try:
        cur = conn.cursor()
        cur.execute('DELETE FROM mx_users WHERE email=%s',(email,))
        conn.commit(); cur.close(); conn.close(); return True
    except Exception as e: print('db_delete:', e); return False
def db_update_role(email, role):
    conn = get_db()
    if not conn: return False
    try:
        cur = conn.cursor()
        cur.execute('UPDATE mx_users SET role=%s WHERE email=%s',(role,email))
        conn.commit(); cur.close(); conn.close(); return True
    except Exception as e: print('db_update_role:', e); return False
def db_update_pw(email, pw_hash):
    conn = get_db()
    if not conn: return False
    try:
        cur = conn.cursor()
        cur.execute('UPDATE mx_users SET password_hash=%s WHERE email=%s',(pw_hash,email))
        conn.commit(); cur.close(); conn.close(); return True
    except Exception as e: print('db_update_pw:', e); return False
try:
    init_db()
    existing = db_get_users()
    if existing is not None and len(existing) == 0:
        db_upsert_user(
            ADMIN_EMAIL,
            'Admin',
            hashlib.sha256(ADMIN_PASSWORD.encode()).hexdigest(),
            'admin'
        )
except Exception as e: print('DB init error:', e)
def hash_pw(p):
    return generate_password_hash(p, method='pbkdf2:sha256:600000')


def verify_pw(stored, password):
    """Accept legacy SHA-256 hashes while new passwords use salted PBKDF2."""
    stored = stored or ''
    if stored.startswith('pbkdf2:') or stored.startswith('scrypt:'):
        try:
            return check_password_hash(stored, password)
        except ValueError:
            return False
    return hmac.compare_digest(stored, hashlib.sha256(password.encode()).hexdigest())


def password_needs_rehash(stored):
    stored = stored or ''
    return not (stored.startswith('pbkdf2:') or stored.startswith('scrypt:'))
def is_admin():
    return session.get('user_role') in ('admin', 'super_admin') or session.get('user_email','').lower() == ADMIN_EMAIL.lower()
def require_admin():
    if not is_admin():
        return jsonify({'error':'Admin gerekli'}), 403
    return None
def _default_admin():
    return [{'email': ADMIN_EMAIL,
             'password_hash': hash_pw(ADMIN_PASSWORD),
             'role':'admin','name':'Admin'}]
def _load_users_from_env():
    import base64
    uj = os.environ.get('USERS_JSON','')
    if uj:
        try:
            u = json.loads(base64.b64decode(uj).decode())
            if u: return u
        except: pass
    return None
def _users_to_json_b64(users):
    import base64
    return base64.b64encode(json.dumps(users, ensure_ascii=False).encode()).decode()
def get_users():
    global _users_cache
    with _users_lock:
        if _users_cache is not None:
            return list(_users_cache)
    pg = db_get_users()
    if pg is not None:
        result = pg if pg else _default_admin()
        with _users_lock:
            _users_cache = list(result)
        return result
    env_users = _load_users_from_env()
    if env_users:
        with _users_lock:
            _users_cache = list(env_users)
        return env_users
    try:
        logs = read_logs()
        if logs.get('users'):
            with _users_lock:
                _users_cache = list(logs['users'])
            return logs['users']
    except: pass
    default = _default_admin()
    with _users_lock:
        _users_cache = list(default)
    return default
def save_users(users):
    global _users_cache
    with _users_lock:
        _users_cache = list(users)
    try:
        current = read_logs()
        current['users'] = users
        write_logs(current)
    except: pass
    return _users_to_json_b64(users)
@app.route('/')
@app.route('/ads')
@app.route('/strateji')
@app.route('/optimizasyon')
@app.route('/hedef-kitle')
@app.route('/tasarim')
@app.route('/raporlar')
@app.route('/gorevler')
@app.route('/olustur')
@app.route('/ayarlar')
@app.route('/kullanicilar')
@app.route('/ai')
@app.route('/google-ads')
@app.route('/seo')
@app.route('/analitik')
@app.route('/kampanyalar')
@app.route('/ai-ajans')
@app.route('/pazaryerleri')
@app.route('/yetki-yonetimi')
@app.route('/bildirim-merkezi')
@app.route('/trendyol')
@app.route('/dosyalar')
@app.route('/ads/<section>')
@app.route('/strateji/<section>')
@app.route('/optimizasyon/<section>')
@app.route('/hedef-kitle/<section>')
@app.route('/tasarim/<section>')
@app.route('/raporlar/<section>')
@app.route('/analitik/<section>')
@app.route('/bildirim-merkezi/<section>')
@app.route('/gorevler/<section>')
@app.route('/olustur/<section>')
@app.route('/ayarlar/<section>')
@app.route('/kullanicilar/<section>')
@app.route('/ai/<section>')
@app.route('/google-ads/<section>')
@app.route('/trendyol/<section>')
@app.route('/dosyalar/<section>')
@app.route('/kampanyalar/<section>')
@app.route('/ai-ajans/<section>')
@app.route('/pazaryerleri/<section>')
@app.route('/yetki-yonetimi/<section>')
@app.route('/seo/<section>')
def home(section=None):
    if not session.get('user_email'):
        return send_from_directory('.', 'login.html') if os.path.exists('login.html') else send_from_directory('.', 'index.html')
    return send_from_directory('.', 'index.html')
@app.route('/login')
def login_page():
    return send_from_directory('.', 'login.html') if os.path.exists('login.html') else send_from_directory('.', 'index.html')
@app.route('/auth/me')
def auth_me():
    if not session.get('user_email'):
        return jsonify({'error':'Giris yapilmamis'}), 401
    return jsonify({'email':session['user_email'],'name':session.get('user_name'),'role':session.get('user_role','admin')})
@app.route('/auth/logout', methods=['POST'])
def auth_logout():
    session.clear()
    return jsonify({'ok':True})
@app.route('/auth/forgot-password', methods=['POST'])
def forgot_password():
    data = request.json or {}
    current = read_logs()
    current.setdefault('resetRequests',[]).insert(0,{
        'email':data.get('email',''), 'message':data.get('message','Şifremi unuttum'),
        'time':datetime.utcnow().isoformat(), 'status':'pending'
    })
    write_logs(current)
    return jsonify({'ok':True})
@app.route('/admin/reset-requests')
def admin_reset_requests():
    r = require_admin()
    if r: return r
    return jsonify(read_logs().get('resetRequests',[]))
@app.route('/admin/reset-requests/<int:idx>/resolve', methods=['POST'])
def resolve_request(idx):
    r = require_admin()
    if r: return r
    current = read_logs()
    reqs = current.get('resetRequests',[])
    if 0<=idx<len(reqs): reqs[idx]['status']='resolved'
    write_logs(current)
    return jsonify({'ok':True})
@app.route('/theme.css')
def serve_theme():
    try:
        return send_from_directory('.', 'theme.css')
    except:
        return ':root{}', 200, {'Content-Type': 'text/css'}
@app.route('/proxy-xml', methods=['GET'])
def proxy_xml():
    url = request.args.get('url','')
    if not url: return 'URL gerekli', 400
    try:
        parsed = urlsplit(url)
        hostname = (parsed.hostname or '').lower().rstrip('.')
        if parsed.scheme not in ('http', 'https') or not hostname or parsed.username or parsed.password:
            return jsonify({'error': 'Geçersiz XML adresi'}), 400
        if parsed.port and parsed.port not in (80, 443):
            return jsonify({'error': 'XML adresinde yalnızca standart HTTP/HTTPS portlarına izin verilir'}), 403
        resolved = {info[4][0] for info in socket.getaddrinfo(hostname, parsed.port or (443 if parsed.scheme == 'https' else 80), type=socket.SOCK_STREAM)}
        if not resolved or any(not ipaddress.ip_address(address).is_global for address in resolved):
            return jsonify({'error': 'Private, local veya ayrılmış ağ adreslerine erişim engellendi'}), 403
        r = requests.get(url, timeout=15, headers={'User-Agent':'Madmext XML Proxy/1.0'}, allow_redirects=False)
        if 300 <= r.status_code < 400:
            return jsonify({'error': 'XML kaynağı yönlendirmelerine izin verilmez'}), 502
        r.raise_for_status()
        return r.text, 200, {'Content-Type':'application/xml; charset=utf-8'}
    except (ValueError, socket.gaierror):
        return jsonify({'error': 'XML adresi çözümlenemedi'}), 400
    except Exception as e:
        return jsonify({'error': 'XML kaynağına erişilemedi'}), 502
@app.route('/modules/<path:filename>')
def serve_module(filename):
    return send_from_directory('modules', filename)
@app.route('/madmext-ads.html')
def serve_old():
    return send_from_directory('.', 'index.html')
@app.route('/ga4', methods=['POST'])
def ga4_proxy():
    try:
        data = request.json
        token = get_ga4_token()
        if not token:
            return jsonify({'error': 'GA4 token alınamadı'})
        report_type = data.get('type', 'runReport')
        url = f'https://analyticsdata.googleapis.com/v1beta/properties/{GA4_PROPERTY_ID}:{report_type}'
        r = requests.post(url,
            headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
            json=data.get('body', {}),
            timeout=15
        )
        return jsonify(r.json())
    except Exception as e:
        return jsonify({'error': str(e)})
@app.route('/logs', methods=['GET'])
def get_logs():
    return jsonify(read_logs())
@app.route('/logs/save', methods=['POST'])
def save_logs():
    data = request.json
    current = read_logs()
    if 'budgetLog' in data: current['budgetLog'] = data['budgetLog']
    if 'taskLog' in data: current['taskLog'] = data['taskLog']
    write_logs(current)
    return jsonify({'ok': True})
@app.route('/logs/action', methods=['POST'])
def log_action():
    data = request.json
    current = read_logs()
    current.setdefault('actionLog', []).insert(0, {**data, 'serverTime': datetime.utcnow().isoformat()})
    current['actionLog'] = current['actionLog'][:500]
    write_logs(current)
    return jsonify({'ok': True})
@app.route('/api', methods=['POST'])
def meta_proxy():
    import json as _json
    data = request.json
    endpoint = data['endpoint']
    raw_params = data.get('params', {})
    method = data.get('method', 'GET')
    url = f"https://graph.facebook.com/v19.0/{endpoint}"
    param_list = [('access_token', META_TOKEN)]
    for k, v in raw_params.items():
        if k == 'action_attribution_windows' and isinstance(v, str):
            try:
                arr = _json.loads(v)
                for item in arr:
                    param_list.append((k + '[]', item))
            except:
                param_list.append((k, v))
        elif k == 'time_range' and isinstance(v, str):
            param_list.append((k, v))
        else:
            param_list.append((k, v))
    if method == 'POST':
        r = requests.post(url, params=param_list, timeout=20)
    else:
        r = requests.get(url, params=param_list, timeout=20)
    result = r.json()
    if method == 'POST' and (result.get('success') or result.get('id')):
        try:
            current = read_logs()
            current.setdefault('actionLog', []).insert(0, {
                'type': 'budget_change', 'endpoint': endpoint,
                'params': {k: v for k, v in data.get('params', {}).items() if k != 'access_token'},
                'serverTime': datetime.utcnow().isoformat()
            })
            current['actionLog'] = current['actionLog'][:500]
            write_logs(current)
        except: pass
    return jsonify(result)

# ── GOOGLE ADS ───────────────────────────────────────────────────────────
def gads_get_token():
    if not GADS_REFRESH_TOKEN or not GADS_CLIENT_ID or not GADS_CLIENT_SECRET:
        return None
    try:
        r = requests.post('https://oauth2.googleapis.com/token', data={
            'client_id': GADS_CLIENT_ID,
            'client_secret': GADS_CLIENT_SECRET,
            'refresh_token': GADS_REFRESH_TOKEN,
            'grant_type': 'refresh_token'
        }, timeout=10)
        data = r.json()
        token = data.get('access_token')
        if not token:
            print(f'Google Ads token hatası: {data}')
        return token
    except Exception as e:
        print(f'Google Ads token hatası: {e}')
        return None

def gads_customer_id():
    return GADS_CUSTOMER_ID.replace('-', '').replace(' ', '')


def gads_numeric_id(value, field_name):
    normalized = str(value or '').strip()
    if not re.fullmatch(r'\d+', normalized):
        raise ValueError(f'{field_name} yalnızca rakamlardan oluşmalı')
    return normalized


def gads_date(value, field_name):
    normalized = str(value or '').strip()
    try:
        datetime.strptime(normalized, '%Y-%m-%d')
    except (TypeError, ValueError):
        raise ValueError(f'{field_name} YYYY-MM-DD formatında olmalı')
    return normalized


def gads_date_condition(data):
    """
    date_range veya date_from/date_to'dan GAQL WHERE koşulu üret.
    Desteklenen preset'ler: TODAY, YESTERDAY, LAST_7_DAYS, LAST_14_DAYS,
    LAST_30_DAYS, THIS_MONTH, LAST_MONTH, LAST_BUSINESS_WEEK, THIS_WEEK_SUN_TODAY
    Custom range: date_from + date_to (YYYY-MM-DD)
    """
    date_from = data.get('date_from', '')
    date_to = data.get('date_to', '')
    if date_from or date_to:
        if not date_from or not date_to:
            raise ValueError('date_from ve date_to birlikte gönderilmeli')
        date_from = gads_date(date_from, 'date_from')
        date_to = gads_date(date_to, 'date_to')
        if date_from > date_to:
            raise ValueError('date_from, date_to değerinden sonra olamaz')
        return f"segments.date BETWEEN '{date_from}' AND '{date_to}'"
    
    date_range = str(data.get('date_range', 'LAST_7_DAYS') or 'LAST_7_DAYS')
    # BETWEEN_ prefix'li string gelirse parse et
    if date_range.startswith('BETWEEN_'):
        parts = date_range.replace('BETWEEN_', '').split('_')
        if len(parts) == 2:
            date_from = gads_date(parts[0], 'date_from')
            date_to = gads_date(parts[1], 'date_to')
            if date_from > date_to:
                raise ValueError('date_from, date_to değerinden sonra olamaz')
            return f"segments.date BETWEEN '{date_from}' AND '{date_to}'"
        raise ValueError('BETWEEN tarih aralığı geçersiz')
    
    # Geçerli GAQL preset'leri
    valid_presets = {
        'TODAY', 'YESTERDAY', 'LAST_7_DAYS', 'LAST_14_DAYS', 'LAST_30_DAYS',
        'THIS_MONTH', 'LAST_MONTH', 'LAST_BUSINESS_WEEK', 'THIS_WEEK_SUN_TODAY',
        'LAST_WEEK_SUN_SAT', 'LAST_WEEK_MON_SUN'
    }
    if date_range.upper() in valid_presets:
        return f"segments.date DURING {date_range.upper()}"
    
    # Bilinmeyen preset — varsayılan
    return "segments.date DURING LAST_7_DAYS"

def gads_campaign_rows(data, http_post=None):
    """Google Ads kampanya verisini doğrulanmış GAQL ile salt-okunur getirir."""
    data = data or {}
    date_cond = gads_date_condition(data)
    campaign_id = data.get('campaign_id')
    campaign_filter = ''
    if campaign_id not in ('', None):
        campaign_id = gads_numeric_id(campaign_id, 'campaign_id')
        campaign_filter = f"AND campaign.id = '{campaign_id}'"
    token = gads_get_token()
    if not token:
        return {
            'error': 'Google Ads token alınamadı. GADS_REFRESH_TOKEN, GADS_CLIENT_ID, GADS_CLIENT_SECRET, GADS_DEVELOPER_TOKEN ve GADS_CUSTOMER_ID değerlerini kontrol edin.',
            'configured': bool(GADS_DEVELOPER_TOKEN and GADS_CUSTOMER_ID)
        }

    query = f"""
        SELECT
          campaign.id,
          campaign.name,
          campaign.status,
          campaign.advertising_channel_type,
          campaign_budget.amount_micros,
          campaign_budget.type,
          metrics.cost_micros,
          metrics.conversions,
          metrics.conversions_value,
          metrics.clicks,
          metrics.impressions,
          metrics.ctr,
          metrics.average_cpc,
          metrics.cost_per_conversion
        FROM campaign
        WHERE {date_cond}
          AND campaign.status != 'REMOVED'
          {campaign_filter}
        ORDER BY metrics.cost_micros DESC
        LIMIT 100
    """

    cid = gads_customer_id()
    url = f'https://googleads.googleapis.com/v19/customers/{cid}/googleAds:search'
    headers = {
        'Authorization': f'Bearer {token}',
        'developer-token': GADS_DEVELOPER_TOKEN,
        'Content-Type': 'application/json'
    }
    if GADS_LOGIN_CUSTOMER_ID:
        headers['login-customer-id'] = GADS_LOGIN_CUSTOMER_ID.replace('-', '')

    try:
        r = (http_post or requests.post)(url, headers=headers, json={'query': query}, timeout=20)
        print(f'Google Ads campaigns URL: {url}')
        print(f'Google Ads campaigns status: {r.status_code}')
        print(f'Google Ads campaigns body: {r.text[:500]}')
        
        # Boş yanıt kontrolü
        if not r.text or not r.text.strip():
            return {'rows': [], 'configured': True}
        
        result = r.json()
        print(f'Google Ads campaigns result keys: {list(result.keys())}')

        if 'error' in result:
            err_msg = result['error'].get('message', 'Google Ads API hatası')
            details = result['error'].get('details', [])
            print(f'Google Ads API hatası: {err_msg}, details: {details}')
            return {'error': err_msg, 'configured': True}

        rows = []
        for row in result.get('results', []):
            camp = row.get('campaign', {})
            budget = row.get('campaignBudget', {})
            metrics = row.get('metrics', {})
            rows.append({
                'id': camp.get('id', ''),
                'name': camp.get('name', ''),
                'status': camp.get('status', ''),
                'type': camp.get('advertisingChannelType', ''),
                'budget_micros': int(budget.get('amountMicros', 0) or 0),
                'budget_type': budget.get('type', ''),
                'cost': round(int(metrics.get('costMicros', 0) or 0) / 1_000_000, 2),
                'conversions': round(float(metrics.get('conversions', 0) or 0), 1),
                'conversions_value': round(float(metrics.get('conversionsValue', 0) or 0), 2),
                'clicks': int(metrics.get('clicks', 0) or 0),
                'impressions': int(metrics.get('impressions', 0) or 0),
                'ctr': round(float(metrics.get('ctr', 0) or 0) * 100, 2),
                'avg_cpc': round(int(metrics.get('averageCpc', 0) or 0) / 1_000_000, 2),
                'cpa': round(int(metrics.get('costPerConversion', 0) or 0) / 1_000_000, 2),
            })
        return {'rows': rows, 'configured': True}
    except Exception as e:
        print(f'Google Ads campaigns exception: {e}')
        return {'error': str(e), 'configured': True}


@app.route('/gads/campaigns', methods=['POST'])
def gads_campaigns():
    try:
        return jsonify(gads_campaign_rows(request.json or {}))
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

@app.route('/gads/adgroups', methods=['POST'])
def gads_adgroups():
    data = request.json or {}
    try:
        date_cond = gads_date_condition(data)
        campaign_id = data.get('campaign_id', '')
        if campaign_id not in ('', None):
            campaign_id = gads_numeric_id(campaign_id, 'campaign_id')
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    token = gads_get_token()
    if not token:
        return jsonify({'error': 'Google Ads yapılandırılmamış.', 'configured': False})

    where_extra = f"AND campaign.id = '{campaign_id}'" if campaign_id else ''

    query = f"""
        SELECT
          ad_group.id, ad_group.name, ad_group.status,
          campaign.id, campaign.name,
          metrics.cost_micros, metrics.conversions,
          metrics.clicks, metrics.impressions, metrics.ctr, metrics.cost_per_conversion
        FROM ad_group
        WHERE {date_cond}
          AND ad_group.status != 'REMOVED'
          {where_extra}
        ORDER BY metrics.cost_micros DESC
        LIMIT 100
    """

    cid = gads_customer_id()
    url = f'https://googleads.googleapis.com/v19/customers/{cid}/googleAds:search'
    headers = {
        'Authorization': f'Bearer {token}',
        'developer-token': GADS_DEVELOPER_TOKEN,
        'Content-Type': 'application/json'
    }
    if GADS_LOGIN_CUSTOMER_ID:
        headers['login-customer-id'] = GADS_LOGIN_CUSTOMER_ID.replace('-', '')

    try:
        r = requests.post(url, headers=headers, json={'query': query}, timeout=20)
        if not r.text or not r.text.strip():
            return jsonify({'rows': [], 'configured': True})
        result = r.json()
        if 'error' in result:
            return jsonify({'error': result['error'].get('message', 'API hatası'), 'configured': True})
        rows = []
        for row in result.get('results', []):
            ag = row.get('adGroup', {})
            camp = row.get('campaign', {})
            metrics = row.get('metrics', {})
            rows.append({
                'id': ag.get('id', ''),
                'name': ag.get('name', ''),
                'status': ag.get('status', ''),
                'campaign_id': camp.get('id', ''),
                'campaign_name': camp.get('name', ''),
                'cost': round(int(metrics.get('costMicros', 0) or 0) / 1_000_000, 2),
                'conversions': round(float(metrics.get('conversions', 0) or 0), 1),
                'clicks': int(metrics.get('clicks', 0) or 0),
                'impressions': int(metrics.get('impressions', 0) or 0),
                'ctr': round(float(metrics.get('ctr', 0) or 0) * 100, 2),
                'cpa': round(int(metrics.get('costPerConversion', 0) or 0) / 1_000_000, 2),
            })
        return jsonify({'rows': rows, 'configured': True})
    except Exception as e:
        return jsonify({'error': str(e), 'configured': True})

@app.route('/gads/budget', methods=['POST'])
def gads_update_budget():
    data = request.json or {}
    try:
        budget_id = gads_numeric_id(data.get('budget_id'), 'budget_id')
        new_amount_tl = float(data.get('amount_tl', 0))
    except (TypeError, ValueError) as exc:
        return jsonify({'error': str(exc) or 'Geçersiz parametre', 'success': False}), 400
    if new_amount_tl <= 0:
        return jsonify({'error': 'Geçersiz parametre', 'success': False}), 400
    token = gads_get_token()
    if not token:
        return jsonify({'error': 'Google Ads yapılandırılmamış.', 'success': False})
    cid = gads_customer_id()
    amount_micros = int(new_amount_tl * 1_000_000)
    patch_url = f'https://googleads.googleapis.com/v19/customers/{cid}/campaignBudgets:mutate'
    headers = {
        'Authorization': f'Bearer {token}',
        'developer-token': GADS_DEVELOPER_TOKEN,
        'Content-Type': 'application/json'
    }
    if GADS_LOGIN_CUSTOMER_ID:
        headers['login-customer-id'] = GADS_LOGIN_CUSTOMER_ID.replace('-', '')
    mutate_body = {
        'operations': [{
            'update': {
                'resourceName': f'customers/{cid}/campaignBudgets/{budget_id}',
                'amountMicros': str(amount_micros)
            },
            'updateMask': 'amountMicros'
        }]
    }
    try:
        r = requests.post(patch_url, headers=headers, json=mutate_body, timeout=15)
        result = r.json()
        if 'error' in result:
            return jsonify({'error': result['error'].get('message', 'API hatası'), 'success': False})
        try:
            current = read_logs()
            current.setdefault('actionLog', []).insert(0, {
                'type': 'gads_budget_change', 'budget_id': budget_id,
                'amount_tl': new_amount_tl, 'serverTime': datetime.utcnow().isoformat()
            })
            current['actionLog'] = current['actionLog'][:500]
            write_logs(current)
        except: pass
        return jsonify({'success': True, 'amount_tl': new_amount_tl})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False})

@app.route('/gads/toggle', methods=['POST'])
def gads_toggle_status():
    data = request.json or {}
    try:
        campaign_id = gads_numeric_id(data.get('campaign_id'), 'campaign_id')
    except ValueError as exc:
        return jsonify({'error': str(exc), 'success': False}), 400
    new_status = data.get('status', 'PAUSED')
    if new_status not in ('ENABLED', 'PAUSED'):
        return jsonify({'error': 'Geçersiz durum', 'success': False}), 400
    token = gads_get_token()
    if not token:
        return jsonify({'error': 'Google Ads yapılandırılmamış.', 'success': False})
    cid = gads_customer_id()
    url = f'https://googleads.googleapis.com/v19/customers/{cid}/campaigns:mutate'
    headers = {
        'Authorization': f'Bearer {token}',
        'developer-token': GADS_DEVELOPER_TOKEN,
        'Content-Type': 'application/json'
    }
    if GADS_LOGIN_CUSTOMER_ID:
        headers['login-customer-id'] = GADS_LOGIN_CUSTOMER_ID.replace('-', '')
    body = {
        'operations': [{
            'update': {
                'resourceName': f'customers/{cid}/campaigns/{campaign_id}',
                'status': new_status
            },
            'updateMask': 'status'
        }]
    }
    try:
        r = requests.post(url, headers=headers, json=body, timeout=15)
        result = r.json()
        if 'error' in result:
            return jsonify({'error': result['error'].get('message', 'API hatası'), 'success': False})
        try:
            current = read_logs()
            current.setdefault('actionLog', []).insert(0, {
                'type': 'gads_status_change', 'campaign_id': campaign_id,
                'status': new_status, 'serverTime': datetime.utcnow().isoformat()
            })
            write_logs(current)
        except: pass
        return jsonify({'success': True, 'status': new_status})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False})

@app.route('/gads/status', methods=['GET'])
def gads_status():
    configured = bool(GADS_DEVELOPER_TOKEN and GADS_CUSTOMER_ID and GADS_CLIENT_ID and GADS_REFRESH_TOKEN)
    if not configured:
        missing = []
        if not GADS_DEVELOPER_TOKEN: missing.append('GADS_DEVELOPER_TOKEN')
        if not GADS_CUSTOMER_ID: missing.append('GADS_CUSTOMER_ID')
        if not GADS_CLIENT_ID: missing.append('GADS_CLIENT_ID')
        if not GADS_REFRESH_TOKEN: missing.append('GADS_REFRESH_TOKEN')
        return jsonify({'configured': False, 'missing': missing})
    token = gads_get_token()
    return jsonify({
        'configured': True,
        'token_ok': bool(token),
        'customer_id': GADS_CUSTOMER_ID,
        'login_customer_id': GADS_LOGIN_CUSTOMER_ID,
        'login_ok': bool(GADS_LOGIN_CUSTOMER_ID)
    })

@app.route('/psi', methods=['GET'])
def psi_proxy():
    """PageSpeed Insights proxy - rate limit korumasi ve key yonetimi"""
    url = request.args.get('url','')
    strategy = request.args.get('strategy','mobile')
    if not url:
        return jsonify({'error':{'message':'url parametresi gerekli'}}), 400
    try:
        api_url = 'https://www.googleapis.com/pagespeedonline/v5/runPagespeed'
        params = {'url': url, 'strategy': strategy, 'locale': 'tr'}
        if PSI_KEY:
            params['key'] = PSI_KEY
        r = requests.get(api_url, params=params, timeout=30)
        # Her durumda JSON don
        try:
            data = r.json()
        except Exception:
            return jsonify({'error':{'message':'PSI API yaniti parse edilemedi: '+r.text[:200]}}), 500
        return jsonify(data), r.status_code
    except requests.Timeout:
        return jsonify({'error':{'message':'PageSpeed API zaman asimi. Tekrar deneyin.'}}), 504
    except Exception as e:
        return jsonify({'error':{'message': str(e)}}), 500

@app.route('/claude/test', methods=['GET'])
def claude_test():
    import json as _json
    key_set = bool(ANTHROPIC_KEY)
    key_len = len(ANTHROPIC_KEY) if ANTHROPIC_KEY else 0
    return _json.dumps({
        'key_set': key_set,
        'key_length': key_len,
        'key_prefix': ANTHROPIC_KEY[:10] + '...' if key_set else 'YOK'
    }), 200, {'Content-Type': 'application/json'}

@app.route('/claude', methods=['POST'])
def claude_proxy():
    # Her zaman JSON don - hic bir kosulda HTML donme
    try:
        data = request.get_json(force=True, silent=True) or {}
        if not ANTHROPIC_KEY:
            app.logger.error('ANTHROPIC_KEY tanimli degil')
            return jsonify({'error': {'message': 'ANTHROPIC_KEY Railway ortam degiskeni tanimli degil.'}}), 500
        r = requests.post(
            'https://api.anthropic.com/v1/messages',
            headers={
                'x-api-key': ANTHROPIC_KEY,
                'anthropic-version': '2023-06-01',
                'content-type': 'application/json'
            },
            json=data,
            timeout=90
        )
        try:
            return jsonify(r.json()), r.status_code
        except Exception:
            return jsonify({'error': {'message': 'Anthropic API yaniti parse edilemedi: ' + r.text[:200]}}), 500
    except requests.Timeout:
        return jsonify({'error': {'message': 'Zaman asimi (90s). Tekrar deneyin.'}}), 504
    except Exception as e:
        app.logger.error('Claude proxy hatasi: ' + str(e))
        return jsonify({'error': {'message': str(e)}}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)

# ── TRENDYOL ─────────────────────────────────────────────────────────────
import io as _io

def ty_init_db():
    conn = get_db()
    if not conn: return False
    try:
        cur = conn.cursor()
        # Ürün reklamları kampanya bazlı
        cur.execute("""CREATE TABLE IF NOT EXISTS ty_urun (
            id SERIAL PRIMARY KEY,
            ad TEXT, statu TEXT, baslangic TEXT, bitis TEXT,
            urun_adedi TEXT, content_ids TEXT,
            toplam_butce TEXT, gunluk_butce TEXT, kalan_butce TEXT,
            harcama TEXT, tbm_teklifi TEXT, gerceklesen_tbm TEXT,
            tiklanma TEXT, goruntulenme TEXT,
            dogrudan_satis TEXT, dolayli_satis TEXT, toplam_satis TEXT,
            dogrudan_ciro TEXT, dolayli_ciro TEXT, toplam_ciro TEXT, roas TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(ad, baslangic)
        )""")
        try:
            cur.execute("ALTER TABLE ty_urun ADD COLUMN IF NOT EXISTS content_ids TEXT")
        except: pass
        # Ürün bazlı detay raporu
        cur.execute("""CREATE TABLE IF NOT EXISTS ty_urun_detay (
            id SERIAL PRIMARY KEY,
            kampanya TEXT, urun_adi TEXT, content_id TEXT, model TEXT,
            harcama TEXT, goruntulenme TEXT, tiklanma TEXT, ctr TEXT,
            dogrudan_satis TEXT, dolayli_satis TEXT, toplam_satis TEXT,
            dogrudan_ciro TEXT, dolayli_ciro TEXT, toplam_ciro TEXT, roas TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(kampanya, content_id)
        )""")
        # Mağaza reklamları
        cur.execute("""CREATE TABLE IF NOT EXISTS ty_magaza (
            id SERIAL PRIMARY KEY,
            ad TEXT, statu TEXT, baslangic TEXT, bitis TEXT,
            harcama TEXT, goruntulenme TEXT, tiklanma TEXT,
            toplam_satis TEXT, toplam_ciro TEXT, roas TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(ad, baslangic)
        )""")
        # Influencer reklamları
        cur.execute("""CREATE TABLE IF NOT EXISTS ty_influencer (
            id SERIAL PRIMARY KEY,
            ad TEXT, statu TEXT, baslangic TEXT, bitis TEXT,
            butce_tipi TEXT, odeme TEXT, ziyaret TEXT,
            satis TEXT, ciro TEXT, paylasim TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(ad, baslangic)
        )""")
        # Meta reklamları (Trendyol üzerinden)
        cur.execute("""CREATE TABLE IF NOT EXISTS ty_meta (
            id SERIAL PRIMARY KEY,
            ad TEXT, statu TEXT, baslangic TEXT, bitis TEXT,
            toplam_butce TEXT, harcama TEXT, goruntulenme TEXT, tiklanma TEXT,
            satis TEXT, ciro TEXT, roas TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(ad, baslangic)
        )""")
        conn.commit(); cur.close(); conn.close()
        return True
    except Exception as e:
        print('ty_init_db:', e)
        try: conn.close()
        except: pass
        return False

try: ty_init_db()
except: pass

def ty_detect(ws):
    """Excel tipini tespit et"""
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    h = set(headers)
    if 'Content Id' in h or 'Model Kodu' in h:
        return 'urun_detay'
    if 'Reklam Adı' in h and 'Harcama Getirisi' in h and 'TBM Teklifi' in h:
        return 'urun'
    if 'Reklam Adı' in h and 'Bütçe Tipi' in h:
        return 'influencer'
    if 'Reklam Adı' in h and 'Toplam Bütçe' in h and 'Reklam Cirosu' in h:
        return 'meta'
    if 'Reklam Adı' in h and 'Harcanan Bütçe' in h:
        return 'magaza'
    # Fallback: sütun sayısına göre
    if ws.max_column >= 20:
        return 'urun'
    return 'unknown'

def safe(v):
    if v is None: return ''
    return str(v).strip()

@app.route('/trendyol/upload', methods=['POST'])
def trendyol_upload():
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({'error': 'Dosya bulunamadı', 'success': False})
        try:
            import openpyxl
        except ImportError:
            return jsonify({'error': 'openpyxl kurulu değil', 'success': False})

        data = file.read()
        wb = openpyxl.load_workbook(_io.BytesIO(data))
        ws = wb.active
        rtype = ty_detect(ws)
        sheet_name = ws.title

        if rtype == 'urun':
            return _ty_upload_urun(ws)
        elif rtype == 'urun_detay':
            return _ty_upload_urun_detay(ws, sheet_name)
        elif rtype == 'magaza':
            return _ty_upload_magaza(ws)
        elif rtype == 'influencer':
            return _ty_upload_influencer(ws)
        elif rtype == 'meta':
            return _ty_upload_meta(ws)
        else:
            return jsonify({'error': 'Tanınmayan dosya. Trendyol Excel raporlarından birini yükleyin.', 'success': False})
    except Exception as e:
        return jsonify({'error': str(e), 'success': False})

def _ty_upsert(conn, table, rows, cols, conflict_cols, update_cols):
    """Genel upsert fonksiyonu"""
    if not rows: return 0
    cur = conn.cursor()
    placeholders = ','.join(['%s']*len(cols))
    col_str = ','.join(cols)
    conflict = ','.join(conflict_cols)
    update = ','.join([c+'=EXCLUDED.'+c for c in update_cols])
    count = 0
    for row in rows:
        try:
            cur.execute(
                f'INSERT INTO {table} ({col_str}) VALUES ({placeholders}) '
                f'ON CONFLICT ({conflict}) DO UPDATE SET {update}',
                row
            )
            count += 1
        except Exception as e:
            print(f'{table} upsert hata:', e)
    conn.commit(); cur.close()
    return count

def _ty_upload_urun(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        rows.append((
            safe(row[0]), safe(row[1]), safe(row[2]), safe(row[3]),
            safe(row[4]), safe(row[5]),  # content_ids
            safe(row[6]), safe(row[7]), safe(row[8]),
            safe(row[9]), safe(row[10]), safe(row[11]),
            safe(row[12]), safe(row[13]),
            safe(row[14]), safe(row[15]), safe(row[16]),
            safe(row[17]), safe(row[18]), safe(row[19]), safe(row[20]),
        ))
    cols = ['ad','statu','baslangic','bitis','urun_adedi','content_ids',
            'toplam_butce','gunluk_butce','kalan_butce','harcama','tbm_teklifi','gerceklesen_tbm',
            'tiklanma','goruntulenme','dogrudan_satis','dolayli_satis','toplam_satis',
            'dogrudan_ciro','dolayli_ciro','toplam_ciro','roas']
    update_cols = ['statu','harcama','tiklanma','goruntulenme','toplam_satis','toplam_ciro','roas','kalan_butce','content_ids']
    conn = get_db()
    if not conn:
        return jsonify({'error': 'Veritabanı bağlantısı yok. DATABASE_URL kontrol edin.', 'success': False})
    try:
        count = _ty_upsert(conn, 'ty_urun', rows, cols, ['ad','baslangic'], update_cols)
        conn.close()
        return jsonify({'success':True,'count':count,'type':'Ürün Reklamları','tab':'urun'})
    except Exception as e:
        return jsonify({'error':str(e),'success':False})

def _ty_upload_urun_detay(ws, kampanya):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        rows.append((
            kampanya, safe(row[0]), safe(row[1]), safe(row[2]),
            safe(row[3]), safe(row[4]), safe(row[5]), safe(row[6]),
            safe(row[7]), safe(row[8]), safe(row[9]),
            safe(row[10]), safe(row[11]), safe(row[12]), safe(row[13]),
        ))
    cols = ['kampanya','urun_adi','content_id','model','harcama','goruntulenme','tiklanma',
            'ctr','dogrudan_satis','dolayli_satis','toplam_satis','dogrudan_ciro',
            'dolayli_ciro','toplam_ciro','roas']
    update_cols = ['harcama','tiklanma','toplam_satis','toplam_ciro','roas']
    conn = get_db()
    if conn:
        try:
            count = _ty_upsert(conn, 'ty_urun_detay', rows, cols, ['kampanya','content_id'], update_cols)
            conn.close()
            return jsonify({'success':True,'count':count,'type':'Ürün Detay ('+kampanya+')','tab':'urun'})
        except Exception as e:
            return jsonify({'error':str(e),'success':False})
    else:
        current = read_logs()
        current.setdefault('ty_urun_detay', [])
        for row in rows:
            current['ty_urun_detay'].append(dict(zip(cols, row)))
        write_logs(current)
        return jsonify({'success':True,'count':len(rows),'type':'Ürün Detay','tab':'urun'})

def _ty_upload_magaza(ws):
    # Mağaza reklam sütunları: Ad, Statü, Başlangıç, Bitiş, Harcama, Gösterim, Tıklanma, Satış, Ciro, ROAS
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        # Sütun sayısına göre esnek parse
        r = list(row) + [''] * 20
        rows.append((safe(r[0]),safe(r[1]),safe(r[2]),safe(r[3]),
                     safe(r[4]),safe(r[5]),safe(r[6]),safe(r[7]),safe(r[8]),safe(r[9])))
    cols = ['ad','statu','baslangic','bitis','harcama','goruntulenme','tiklanma',
            'toplam_satis','toplam_ciro','roas']
    conn = get_db()
    if not conn:
        return jsonify({'error': 'Veritabanı bağlantısı yok.', 'success': False})
    try:
        count = _ty_upsert(conn, 'ty_magaza', rows, cols, ['ad','baslangic'],
                           ['statu','harcama','tiklanma','toplam_satis','toplam_ciro','roas'])
        conn.close()
        return jsonify({'success':True,'count':count,'type':'Mağaza Reklamları','tab':'magaza'})
    except Exception as e:
        return jsonify({'error':str(e),'success':False})

def _ty_upload_influencer(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        r = list(row) + [''] * 20
        rows.append((safe(r[0]),safe(r[1]),safe(r[2]),safe(r[3]),
                     safe(r[4]),safe(r[5]),safe(r[6]),safe(r[7]),safe(r[8]),safe(r[9])))
    cols = ['ad','statu','baslangic','bitis','butce_tipi','odeme','ziyaret',
            'satis','ciro','paylasim']
    conn = get_db()
    if not conn:
        return jsonify({'error': 'Veritabanı bağlantısı yok.', 'success': False})
    try:
        count = _ty_upsert(conn, 'ty_influencer', rows, cols, ['ad','baslangic'],
                           ['statu','odeme','ziyaret','satis','ciro'])
        conn.close()
        return jsonify({'success':True,'count':count,'type':'Influencer Reklamları','tab':'influencer'})
    except Exception as e:
        return jsonify({'error':str(e),'success':False})

def _ty_upload_meta(ws):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        r = list(row) + [''] * 20
        rows.append((safe(r[0]),safe(r[1]),safe(r[2]),safe(r[3]),
                     safe(r[4]),safe(r[5]),safe(r[6]),safe(r[7]),safe(r[8]),safe(r[9]),safe(r[10])))
    cols = ['ad','statu','baslangic','bitis','toplam_butce','harcama',
            'goruntulenme','tiklanma','satis','ciro','roas']
    conn = get_db()
    if not conn:
        return jsonify({'error': 'Veritabanı bağlantısı yok.', 'success': False})
    try:
        count = _ty_upsert(conn, 'ty_meta', rows, cols, ['ad','baslangic'],
                           ['statu','harcama','tiklanma','satis','ciro','roas'])
        conn.close()
        return jsonify({'success':True,'count':count,'type':'Meta Reklamları','tab':'meta'})
    except Exception as e:
        return jsonify({'error':str(e),'success':False})

@app.route('/trendyol/data', methods=['GET'])
def trendyol_data():
    result = {'urun':[],'magaza':[],'influencer':[],'meta':[],'urun_detay':[]}
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            tables = {
                'urun': ('ty_urun', ['ad','statu','baslangic','bitis','urun_adedi','content_ids',
                    'toplam_butce','gunluk_butce','kalan_butce','harcama','tbm_teklifi','gerceklesen_tbm',
                    'tiklanma','goruntulenme','dogrudan_satis','dolayli_satis','toplam_satis',
                    'dogrudan_ciro','dolayli_ciro','toplam_ciro','roas']),
                'magaza': ('ty_magaza', ['ad','statu','baslangic','bitis','harcama',
                    'goruntulenme','tiklanma','toplam_satis','toplam_ciro','roas']),
                'influencer': ('ty_influencer', ['ad','statu','baslangic','bitis',
                    'butce_tipi','odeme','ziyaret','satis','ciro','paylasim']),
                'meta': ('ty_meta', ['ad','statu','baslangic','bitis','toplam_butce',
                    'harcama','goruntulenme','tiklanma','satis','ciro','roas']),
                'urun_detay': ('ty_urun_detay', ['kampanya','urun_adi','content_id','model',
                    'harcama','goruntulenme','tiklanma','ctr','dogrudan_satis','dolayli_satis',
                    'toplam_satis','dogrudan_ciro','dolayli_ciro','toplam_ciro','roas']),
            }
            for key, (table, cols) in tables.items():
                try:
                    cur.execute(f'SELECT {",".join(cols)} FROM {table} ORDER BY baslangic DESC LIMIT 3000' if 'baslangic' in cols else f'SELECT {",".join(cols)} FROM {table} ORDER BY created_at DESC LIMIT 5000')
                    result[key] = [dict(zip(cols, row)) for row in cur.fetchall()]
                except Exception as e:
                    print(f'{table} select:', e)
            cur.close(); conn.close()
        except Exception as e:
            print('trendyol_data:', e)
    return jsonify(result)

@app.route('/trendyol/stats', methods=['GET'])
def trendyol_stats():
    """Özet istatistikler"""
    conn = get_db()
    stats = {}
    if conn:
        try:
            cur = conn.cursor()
            for table, label in [('ty_urun','urun'),('ty_magaza','magaza'),('ty_influencer','influencer'),('ty_meta','meta')]:
                try:
                    cur.execute(f'SELECT COUNT(*) FROM {table}')
                    stats[label] = cur.fetchone()[0]
                except: stats[label] = 0
            cur.close(); conn.close()
        except: pass
    return jsonify(stats)
